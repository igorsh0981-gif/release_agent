"""
PM Agent — Project Manager
Формирует PM артефакты и нарезку задач для projekt (CSV + XLSX).
"""

import io
import json
import logging
import re
from models.task import Task
from services.claude_client import call_claude
from services.notifier import notify_pm_done

logger = logging.getLogger(__name__)

PM_SYSTEM = """Ты — опытный PM банковского мобильного приложения.

На вход получаешь BA, SA и QATC артефакты. Твоя задача: сформировать полный пакет PM артефактов.

Верни ТОЛЬКО JSON (без markdown-обёртки) со следующими ключами:

{
  "summary": "краткое резюме для уведомления (2-3 предложения)",
  "protocol": "## PM Protocol\\n\\n### Цели релиза\\n...\\n### Команда\\n...\\n### Риски\\n...",
  "jira": "## Jira Epics & Stories\\n\\nEpic: [название]\\n  Story: [название]\\n    Task: [название]\\n...",
  "epics": "## Epics\\n\\n### Epic 1: [название]\\nОписание: ...\\nКритерии готовности: ...\\n...",
  "risks": "## Реестр рисков\\n\\n| Риск | Вероятность | Влияние | Митигация |\\n|---|---|---|---|\\n...",
  "raci": "## RACI матрица\\n\\n| Активность | PM | Dev | QA | BA | SA |\\n|---|---|---|---|---|---|\\n...",
  "team": "## Состав команды\\n\\n### Роли и ответственность\\n...",
  "projekt_tasks": [
    {
      "name": "Релиз v{version} - {feature}",
      "duration": 0,
      "level": 0,
      "type": "release",
      "depends_on": ""
    },
    {
      "name": "Обновление БП релиза",
      "duration": 2.5,
      "level": 1,
      "type": "process",
      "depends_on": ""
    },
    {
      "name": "Дизайн фичи",
      "duration": 3,
      "level": 1,
      "type": "design",
      "depends_on": "Обновление БП релиза"
    },
    {
      "name": "Системный анализ",
      "duration": 12,
      "level": 1,
      "type": "sa",
      "depends_on": "Дизайн фичи"
    },
    {
      "name": "Разработка БЭК",
      "duration": 30,
      "level": 1,
      "type": "dev_back",
      "depends_on": "Системный анализ"
    },
    {
      "name": "Разработка Фронт",
      "duration": 30,
      "level": 1,
      "type": "dev_front",
      "depends_on": "Разработка БЭК"
    },
    {
      "name": "Тестирование",
      "duration": 20,
      "level": 1,
      "type": "qa",
      "depends_on": "Разработка Фронт"
    },
    {
      "name": "Демонстрация функционала",
      "duration": 1,
      "level": 1,
      "type": "demo",
      "depends_on": "Тестирование"
    },
    {
      "name": "Формирование документации",
      "duration": 0.25,
      "level": 1,
      "type": "docs",
      "depends_on": "Демонстрация функционала"
    },
    {
      "name": "Сопровождение релиза",
      "duration": 150,
      "level": 1,
      "type": "support",
      "depends_on": "Демонстрация функционала"
    }
  ]
}

## Важно:
- Длительности задач бери из SA (разработка) и QATC (тестирование) артефактов
- version определи из контекста или используй "X.X.X"
- projekt_tasks — это нарезка для импорта в систему управления проектами
- level 0 = корневая задача (релиз), level 1 = подзадача
- Пиши на русском языке
"""


async def run_pm(task: Task, bot, notify_chat_id: int) -> Task:
    """
    Запускает PM агента.
    Возвращает task с заполненными всеми PM артефактами.
    """
    logger.info(f"PM старт | задача: {task.feature_name}")

    user_text = (
        f"## Задача: {task.feature_name}\\n"
        f"## BA Артефакт:\\n{task.ba_text}\\n\\n"
        f"## SA Артефакт:\\n{task.sa_text}\\n\\n"
        f"## QATC Артефакт:\\n{task.qatc_text}"
    )

    try:
        response = await call_claude(PM_SYSTEM, user_text, max_tokens=8192, timeout=180)

        # Парсим JSON ответ
        clean = response.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        data = json.loads(clean)

        task.pm_summary = data.get("summary", "")
        task.pm_protocol = data.get("protocol", "")
        task.pm_jira = data.get("jira", "")
        task.pm_epics = data.get("epics", "")
        task.pm_risks = data.get("risks", "")
        task.pm_raci = data.get("raci", "")
        task.pm_team = data.get("team", "")

        # Генерируем CSV и XLSX из projekt_tasks
        projekt_tasks = data.get("projekt_tasks", [])
        task.pm_projekt_csv = _build_csv(projekt_tasks, task.feature_name)
        task.pm_projekt_xlsx = _build_xlsx(projekt_tasks, task.feature_name)

        await notify_pm_done(bot, notify_chat_id, task.pm_summary)
        logger.info("PM завершён")
        return task

    except json.JSONDecodeError as e:
        logger.error(f"PM JSON parse error: {e}")
        # Fallback — сохраняем как текст
        task.pm_protocol = response
        task.pm_summary = response[:200]
        task.pm_projekt_csv = _build_default_csv(task.feature_name)
        task.pm_projekt_xlsx = _build_xlsx(_default_tasks(task.feature_name), task.feature_name)
        await notify_pm_done(bot, notify_chat_id, task.pm_summary)
        return task
    except Exception as e:
        logger.error(f"PM ошибка: {e}")
        raise


def _build_csv(tasks: list, feature_name: str) -> str:
    """Формирует CSV для импорта в projekt"""
    lines = ["Название задачи\tДлительность\tТип\tЗависимость"]
    for t in tasks:
        indent = "   " * t.get("level", 0)
        name = f"{indent}{t.get('name', '')}"
        duration = t.get("duration", 0)
        task_type = t.get("type", "")
        depends = t.get("depends_on", "")
        lines.append(f"{name}\t{duration}\t{task_type}\t{depends}")
    return "\n".join(lines)


def _build_xlsx(tasks: list, feature_name: str) -> bytes:
    """Формирует XLSX для импорта в projekt"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projekt Tasks"

        # Заголовки
        headers = ["Название задачи", "Длительность (дни)", "Тип", "Зависимость"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a73e8")
            cell.alignment = Alignment(horizontal="center")

        # Данные
        for row_idx, t in enumerate(tasks, 2):
            indent = "   " * t.get("level", 0)
            name = f"{indent}{t.get('name', '')}"
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=t.get("duration", 0))
            ws.cell(row=row_idx, column=3, value=t.get("type", ""))
            ws.cell(row=row_idx, column=4, value=t.get("depends_on", ""))

            # Строки первого уровня — полужирный
            if t.get("level", 0) == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).font = Font(bold=True)

        # Ширина колонок
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 35

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        logger.warning("openpyxl не установлен — XLSX не сгенерирован")
        return b""


def _default_tasks(feature_name: str) -> list:
    """Дефолтная нарезка если PM не вернул задачи"""
    return [
        {"name": f"Релиз - {feature_name}", "duration": 0, "level": 0, "type": "release", "depends_on": ""},
        {"name": "Обновление БП релиза", "duration": 2.5, "level": 1, "type": "process", "depends_on": ""},
        {"name": "Дизайн фичи", "duration": 3, "level": 1, "type": "design", "depends_on": "Обновление БП релиза"},
        {"name": "Системный анализ", "duration": 12, "level": 1, "type": "sa", "depends_on": "Дизайн фичи"},
        {"name": "Разработка БЭК", "duration": 30, "level": 1, "type": "dev_back", "depends_on": "Системный анализ"},
        {"name": "Разработка Фронт", "duration": 30, "level": 1, "type": "dev_front", "depends_on": "Разработка БЭК"},
        {"name": "Тестирование", "duration": 20, "level": 1, "type": "qa", "depends_on": "Разработка Фронт"},
        {"name": "Демонстрация функционала", "duration": 1, "level": 1, "type": "demo", "depends_on": "Тестирование"},
        {"name": "Формирование документации", "duration": 0.25, "level": 1, "type": "docs", "depends_on": "Демонстрация функционала"},
        {"name": "Сопровождение релиза", "duration": 150, "level": 1, "type": "support", "depends_on": "Демонстрация функционала"},
    ]


def _build_default_csv(feature_name: str) -> str:
    return _build_csv(_default_tasks(feature_name), feature_name)
